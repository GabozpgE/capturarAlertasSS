# -*- coding: utf-8 -*-


import re
import sys
import shutil
import unicodedata
import datetime as dt
from pathlib import Path

import pdfplumber
import openpyxl
from openpyxl.styles import PatternFill

# PyMuPDF (fitz) se usa solo para extraer la foto de la persona del PDF.
# Lo importamos de forma segura: si no esta instalado, la extraccion de
# fotos simplemente no estara disponible, pero el resto del programa
# sigue funcionando igual.
try:
    import fitz  # PyMuPDF
    _HAY_FITZ = True
except Exception:
    _HAY_FITZ = False
from copy import copy

# ----------------- CONFIGURACION ----------------- uwu 
EXCEL_DEFAULT = "BASE ALERTAS NACIONAL 2026 (JOAO).xlsx"
HOJA_DEFAULT = "Hoja1"          # se usa solo si no se especifica otra
ANIO_EXPEDIENTE = 2026          
PRIMERA_FILA_DATOS = 2           # los datos empiezan en la fila 2

# Las columnas NO se fijan por posicion. Se detectan leyendo los
# encabezados de la fila 1. Asi el Excel puede reordenar o renombrar columnas sin romper
# el programa.



class ErrorPDF(Exception):
    """Error con mensaje claro al procesar un PDF (para mostrar al usuario)."""
    pass


def leer_texto_pdf(ruta_pdf: str) -> str:
    """Extrae todo el texto del PDF en una sola cadena.

    Lanza ErrorPDF con un mensaje entendible si el archivo esta dañado,
    protegido con contraseña, o no es un PDF valido.
    """
    try:
        partes = []
        with pdfplumber.open(ruta_pdf) as pdf:
            if len(pdf.pages) == 0:
                raise ErrorPDF("el PDF no tiene paginas")
            for pagina in pdf.pages:
                txt = pagina.extract_text() or ""
                partes.append(txt)
        texto = "\n".join(partes)
        if not texto.strip():
            raise ErrorPDF("el PDF no contiene texto legible "
                           "(podria ser una imagen escaneada)")
        return texto
    except ErrorPDF:
        raise
    except Exception as e:
        # pdfplumber/pdfminer lanzan varios tipos; los unificamos
        raise ErrorPDF(f"no se pudo leer el PDF (archivo dañado o "
                       f"protegido): {type(e).__name__}")


def extraer_foto_persona(ruta_pdf: str):
    """
    Saca la foto de la persona desaparecida del PDF y devuelve sus bytes
    (en el formato original, normalmente JPEG), o None si no la encuentra.

    El PDF trae varias imagenes (logos, escudos, banners y la foto). Para
    quedarnos con la FOTO y no con un logo, la distinguimos por dos rasgos
    que solo cumple un retrato:
      - proporcion vertical: mas alta que ancha (los logos son anchos/planos)
      - tamaño considerable: no un iconito de pocos pixeles

    Detalle importante: algunos PDF guardan la foto JUNTO con una "mascara"
    del mismo tamaño (para las esquinas redondeadas del marco). Esa mascara
    es casi toda de un color, asi que pesa poquisimos bytes y, si la eligieramos,
    la imagen saldria en blanco. Por eso entre las candidatas nos quedamos con
    la que MAS BYTES tiene: esa es la foto real con detalle, no la mascara.
    """
    if not _HAY_FITZ:
        return None
    try:
        doc = fitz.open(ruta_pdf)
    except Exception:
        return None

    mejor = None          # (bytes_peso, datos_imagen, extension)
    try:
        for pagina in doc:
            for img in pagina.get_images(full=True):
                xref = img[0]
                try:
                    base = doc.extract_image(xref)
                except Exception:
                    continue
                w = base.get("width", 0)
                h = base.get("height", 0)
                if w == 0 or h == 0:
                    continue
                ratio = h / w              # >1 = mas alta que ancha (retrato)
                area = w * h
                peso = len(base["image"])  # bytes: distingue foto de mascara
                # Filtro: forma de retrato y tamaño minimo razonable.
                es_retrato = ratio >= 1.15
                es_grande = area >= 150 * 150
                if es_retrato and es_grande:
                    # Nos quedamos con la de mayor PESO (la foto real, no la
                    # mascara blanca que pesa casi nada).
                    if mejor is None or peso > mejor[0]:
                        mejor = (peso, base["image"], base["ext"])
    finally:
        doc.close()

    return (mejor[1], mejor[2]) if mejor else None


def guardar_foto(ruta_pdf: str, fub: str, carpeta_destino, log=print) -> bool:
    """
    Extrae la foto del PDF y la guarda como {FUB}.jpg en 'carpeta_destino'.
    Devuelve True si la guardo, False si no habia foto o algo fallo.

    El FUB se limpia de caracteres que no sirven para nombre de archivo.
    """
    resultado = extraer_foto_persona(ruta_pdf)
    if resultado is None:
        return False
    datos_img, _ext = resultado

    # Limpiar el FUB para que sea un nombre de archivo valido (sin / \ : etc.)
    nombre = re.sub(r'[\\/:*?"<>|]', "_", (fub or "").strip())
    if not nombre:
        nombre = "sin_fub"

    carpeta_destino = Path(carpeta_destino)
    try:
        carpeta_destino.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        log(f"      (no se pudo crear la carpeta de fotos: {e})")
        return False

    destino = carpeta_destino / f"{nombre}.jpg"
    try:
        # Guardamos los bytes tal cual. La foto viene en JPEG dentro del PDF,
        # asi que escribir con extension .jpg es directo y sin reconvertir.
        with open(destino, "wb") as f:
            f.write(datos_img)
        return True
    except Exception as e:
        log(f"      (no se pudo guardar la foto de {nombre}: {e})")
        return False


def buscar(patron: str, texto: str, grupo: int = 1, flags=re.IGNORECASE) -> str:
    """Aplica una regex y devuelve el grupo limpio, o '' si no encuentra."""
    m = re.search(patron, texto, flags)
    if not m:
        return ""
    return m.group(grupo).strip()


def fecha_a_objeto(fecha_str: str):
    """
    Convierte 'DD/MM/YYYY' (o 'DD-MM-YYYY') a un objeto date real de Python.

    Devolver una fecha real (no texto) es importante: la celda del Excel ya
    tiene su propio formato de fecha (DD-MM-AA), asi que Excel la mostrara
    correctamente y NO marcara el triangulo verde de "fecha como texto".

    Si el texto no tiene forma de fecha, devuelve la cadena original en
    mayusculas (para no perder informacion en casos raros uwu).
    """
    fecha_str = (fecha_str or "").strip()
    m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", fecha_str)
    if not m:
        return fecha_str.upper()
    d, mth, y = m.groups()
    y = int(y)
    if y < 100:               # año de 2 digitos -> asumir 2000+
        y += 2000
    try:
        return dt.date(y, int(mth), int(d))
    except ValueError:
        # fecha imposible (ej. 31/02); devuelve el texto para no romper
        return fecha_str.upper()


def normalizar_lugar(lugar_raw: str) -> str:
    """
    Del PDF 'ESTADO: CIUDAD DE MEXICO , MUNICIPIO: IZTAPALAPA'
    extrae solo el estado.
    """
    m = re.search(r"ESTADO:\s*(.+?)\s*,\s*MUNICIPIO", lugar_raw, re.IGNORECASE)
    if m:
        return m.group(1).strip().upper()
    return lugar_raw.strip().upper()


def _normalizar_encabezado(texto) -> str:
    """
    Normaliza un texto de encabezado para poder compararlo sin importar
    acentos, mayusculas, saltos de linea ni signos.
    'FECHA \\nDE HECHO' y 'fecha de hecho' quedan iguales.
    'Características físicas' y 'CARACTERISTICAS FISICAS' quedan iguales.
    """
    # El problema que resuelve esto: los encabezados del Excel vienen con
    # acentos, saltos de linea metidos a mano ('FECHA \n DE HECHO') y mayusculas
    # inconsistentes. Si comparamos tal cual, nunca coinciden. La solucion es
    # "aplanar" ambos lados a una forma comun antes de comparar.
    if texto is None:
        return ""
    t = str(texto)
    # NFD separa cada letra de su acento en dos caracteres; luego tiramos los
    # acentos sueltos (categoria "Mn"). Resultado: texto sin tildes.
    t = unicodedata.normalize("NFD", t)
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    # todo a minusculas, cambiamos cualquier signo raro por espacio, y
    # colapsamos espacios repetidos. Queda una version "limpia" y comparable.
    t = t.lower()
    t = re.sub(r"[^a-z0-9ñ ]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


# Para cada dato interno, la lista de encabezados posibles (como aparecen
# en el Excel). El programa busca cualquiera de estos nombres en la fila 1.
# Asi, si un dia se reordenan o renombran ligeramente las columnas, sigue
# funcionando mientras el encabezado sea reconocible.
MAPA_ENCABEZADOS = {
    "expediente":            ["expediente"],
    "fecha_registro":        ["fecha de registro"],
    "nombre":                ["nombre"],
    "edad_desaparecer":      ["edad al desaparecer",
                              "edad al momento de la desaparicion"],
    "edad_actual":           ["edad actual"],
    "sexo":                  ["sexo"],
    "genero":                ["genero"],
    "nacionalidad":          ["nacionalidad"],
    "fub":                   ["fub", "folio unico de identificacion"],
    "fecha_hecho":           ["fecha de hecho", "fecha de hechos"],
    "hora_hecho":            ["hora de hecho", "hora de hechos"],
    "fecha_percato":         ["fecha de percato"],
    "hora_percato":          ["hora de percato"],
    "lugar_hechos":          ["lugar de hechos", "lugar de los hechos"],
    "municipio":             ["municipio"],
    "lugar_nacimiento":      ["lugar de nacimiento"],
    "habla_espanol":         ["habla espanol"],
    "lengua_indigena":       ["idioma o lengua indigena", "lengua indigena"],
    "discapacidad":          ["discapacidad"],
    "carpeta_investigacion": ["carpeta de investigacion"],
    "caracteristicas_fisicas": ["caracteristicas fisicas"],
    "senas_particulares":    ["senas particulares"],
    "prendas_vestir":        ["prendas de vestir"],
    "autoridad":             ["autoridad que ingreso el reporte",
                              "autoridades competentes", "autoridad"],
    # 'id' no se mapea: se respeta el que ya viene en la columna A.
}


def mapear_columnas(ws, mapa_encabezados=None) -> dict:
    """
    Lee la fila 1 (encabezados) y devuelve un diccionario que dice, para
    cada dato interno, en que numero de columna esta.
    Ejemplo: {'nombre': 4, 'fub': 10, 'municipio': 16, ...}

    'mapa_encabezados' es {clave: [encabezados posibles]}. Normalmente viene
    de la configuracion del usuario. Si no se pasa, usa el mapa por defecto
    fijo en el codigo (MAPA_ENCABEZADOS), como respaldo.

    Si un encabezado no aparece en la hoja, ese dato simplemente no se
    escribe (la hoja quiza no tiene esa columna).
    """
    if mapa_encabezados is None:
        mapa_encabezados = MAPA_ENCABEZADOS

    # Primer paso: recorremos la fila 1 del Excel y armamos un diccionario
    # {encabezado_normalizado: numero_de_columna}. Es como hacer un indice de
    # "en que columna esta cada titulo" para luego buscar rapido.
    encabezados = {}
    for col in range(1, ws.max_column + 1):
        norm = _normalizar_encabezado(ws.cell(row=1, column=col).value)
        if norm:
            encabezados[norm] = col

    # Segundo paso: por cada dato que queremos escribir, probamos sus nombres
    # posibles (normalizados) contra el indice de arriba. El primero que exista
    # en la hoja, esa es su columna. Si ninguno aparece, ese dato no se mapea y
    # simplemente no se escribira (la hoja no tiene esa columna, y ya).
    mapa = {}
    for dato, posibles in mapa_encabezados.items():
        for nombre in posibles:
            clave = _normalizar_encabezado(nombre)
            if clave in encabezados:
                mapa[dato] = encabezados[clave]
                break
    return mapa


def buscar_columna_id(ws):
    """Devuelve el numero de columna donde esta el encabezado 'ID'."""
    for col in range(1, ws.max_column + 1):
        if _normalizar_encabezado(ws.cell(row=1, column=col).value) == "id":
            return col
    return 1  # por defecto la A


def extraer_estado_municipio(lugar_raw: str):
    """
    Del PDF 'ESTADO: SONORA , MUNICIPIO: HERMOSILLO' devuelve una tupla
    (estado, municipio). Si no encuentra el municipio, devuelve cadena vacia.
    """
    estado = ""
    municipio = ""
    m_est = re.search(r"ESTADO:\s*(.+?)\s*,\s*MUNICIPIO", lugar_raw, re.IGNORECASE)
    if m_est:
        estado = m_est.group(1).strip().upper()
    else:
        estado = lugar_raw.strip().upper()
    m_mun = re.search(r"MUNICIPIO:\s*(.+)$", lugar_raw, re.IGNORECASE)
    if m_mun:
        municipio = m_mun.group(1).strip().upper()
    return estado, municipio


def extraer_bloque_datos(texto: str, etiqueta_ini: str, etiqueta_fin: str) -> str:
    """
    Extrae un bloque de la seccion DATOS (caracteristicas fisicas, señas
    particulares, prendas de vestir) como un texto completo.

    El texto de estos bloques puede venir partido en varias lineas y con la
    etiqueta intercalada, asi que capturamos todo entre la etiqueta inicial
    y la siguiente etiqueta, y limpiamos saltos de linea.
    """
    patron = (re.escape(etiqueta_ini) + r"\s*(.*?)\s*" +
              r"(?=" + etiqueta_fin + r")")
    m = re.search(patron, texto, re.IGNORECASE | re.DOTALL)
    if not m:
        return ""
    bloque = m.group(1)
    # Unir lineas y quitar la propia etiqueta si quedo intercalada
    bloque = re.sub(r"\s*\n\s*", " ", bloque)
    bloque = re.sub(re.escape(etiqueta_ini), " ", bloque, flags=re.IGNORECASE)
    bloque = re.sub(r"\s+", " ", bloque).strip()
    return bloque


def extraer_campo_simple(texto: str, etiquetas: list,
                         otras_etiquetas=None) -> str:
    """
    Extrae un campo del tipo 'etiqueta: valor'.

    Como en el PDF varios campos comparten linea
    (ej. 'Sexo: HOMBRE Genero: MASCULINO'), el valor se corta cuando
    aparece CUALQUIER otra etiqueta conocida, o al fin de linea.

    - etiquetas: las etiquetas posibles de ESTE campo (se prueba cada una).
    - otras_etiquetas: lista de etiquetas de OTROS campos, para saber donde
      cortar. Si no se pasa, corta en fin de linea.
    """
    otras = otras_etiquetas or []
    for etq in etiquetas:
        etq = etq.strip()
        if not etq:
            continue
        m = re.search(re.escape(etq), texto, re.IGNORECASE)
        if not m:
            continue
        # Agarramos todo lo que sigue despues de la etiqueta, pero solo hasta
        # el fin de la linea. El valor de un campo nunca cruza a otra linea.
        inicio = m.end()
        resto = texto[inicio:]
        fin_linea = resto.find("\n")
        if fin_linea != -1:
            resto = resto[:fin_linea]
        # Aqui esta el truco fino: como en el PDF varias etiquetas comparten la
        # misma linea ('Sexo: MUJER Genero: FEMENINO'), buscamos donde empieza
        # la SIGUIENTE etiqueta conocida y cortamos justo antes. Nos quedamos
        # con la posicion de corte mas temprana que encontremos.
        corte = len(resto)
        for otra in otras:
            otra = otra.strip()
            if not otra:
                continue
            mo = re.search(re.escape(otra), resto, re.IGNORECASE)
            if mo and mo.start() < corte:
                corte = mo.start()
        # strip(" :") por si quedo un ": " colgando al inicio o final.
        return resto[:corte].strip(" :")
    return ""


def extraer_datos(texto: str, config=None) -> dict:
    """
    Extrae cada campo del flyer.

    Los campos ESPECIALES (nombre, fub, lugar, fechas, caracteristicas,
    etc.) tienen logica propia y se extraen siempre igual.

    Los campos SIMPLES (etiqueta: valor) se extraen usando las etiquetas
    de la configuracion del usuario si se pasa 'config'; si no, se usan
    las etiquetas por defecto del codigo.
    """
    d = {}

    # Toma el nombre del titulo de la ficha (suele venir bien escrito).
    # Esta entre 'FICHA DE BUSQUEDA DE PERSONA DESAPARECIDA' y
    # 'Folio Unico de Identificacion'.
    d["nombre"] = buscar(
        r"FICHA DE B[UÚ]SQUEDA DE PERSONA DESAPARECIDA\s*(.+?)\s*"
        r"Folio [UÚ]nico de Identificaci[oó]n",
        texto, flags=re.IGNORECASE | re.DOTALL)
    # Respaldo: si el titulo no se encuentra, usa el campo Nombre Social.
    if not d["nombre"]:
        d["nombre"] = buscar(
            r"Nombre Social:\s*(.+?)\s*(?:Nacionalidad:|\n)", texto)

    d["edad_desaparecer"] = buscar(
        r"Edad al momento de la desaparici[oó]n:\s*(\d+)", texto)
    d["edad_actual"] = buscar(r"Edad Actual:\s*(\d+)", texto)

    d["fub"] = buscar(
        r"Folio [UÚ]nico de Identificaci[oó]n\s*([A-Z0-9\-]+)", texto)

    fh = buscar(r"Fecha de hechos:\s*([\d/]+)", texto)
    fp = buscar(r"Fecha de percato:\s*([\d/]+)", texto)
    d["fecha_hecho"] = fecha_a_objeto(fh)
    d["fecha_percato"] = fecha_a_objeto(fp)

    lugar = buscar(r"Lugar de los hechos:\s*(.+?)\s*(?:Carpeta|\n)", texto)
    estado, municipio = extraer_estado_municipio(lugar)
    d["lugar_hechos"] = estado
    d["municipio"] = municipio

    # ---------- Campos SIMPLES (configurables) ----------
    # Si hay config, usamos sus etiquetas; si no, las etiquetas por defecto.
    if config is not None:
        try:
            import config_campos
            simples = config_campos.campos_simples_desde_config(config)
        except Exception:
            simples = {}
    else:
        simples = {}

    # Etiquetas por defecto de los campos simples (respaldo si no hay config)
    simples_default = {
        "sexo": ["Sexo:"],
        "genero": ["Genero:", "Género:"],
        "nacionalidad": ["Nacionalidad:"],
        "lugar_nacimiento": ["Lugar de nacimiento:"],
        "habla_espanol": ["¿Habla español?:", "Habla espanol?:"],
        "lengua_indigena": ["Idioma o lengua indígena:",
                            "Idioma o lengua indigena:"],
        "discapacidad": ["Discapacidad:"],
        "hora_hecho": ["Hora de hechos:"],
        "hora_percato": ["Hora de percato:"],
        "carpeta_investigacion": ["Carpeta de investigación",
                                 "Carpeta de investigacion"],
        "autoridad": ["Autoridades Competentes:"],
    }

    # Reunir TODAS las etiquetas conocidas (para saber donde cortar cada valor)
    todas_las_etiquetas = []
    for etqs in simples_default.values():
        todas_las_etiquetas.extend(etqs)
    # etiquetas de campos que comparten linea y no estan en simples_default
    todas_las_etiquetas.extend(["Edad al momento de la desaparicion:",
                                "Edad Actual:", "Genero:", "Género:",
                                "Nombre Social:", "Idioma", "Fecha de hechos:",
                                "Fecha de percato:", "Hora de hechos:",
                                "Hora de percato:"])

    for clave, etqs_default in simples_default.items():
        etiquetas = simples.get(clave, etqs_default)
        # las "otras" son todas menos las de este campo
        otras = [e for e in todas_las_etiquetas if e not in etiquetas]
        d[clave] = extraer_campo_simple(texto, etiquetas, otras)

    # ---------- Campos ESPECIALES (logica propia, no configurable) ----------
    # El caso raro de todo el programa: en el PDF la etiqueta 'Características
    # físicas:' NO viene al inicio de su texto, sino EN MEDIO. O sea, parte del
    # texto va antes de la etiqueta y parte despues. Por eso no lo podemos leer
    # como un 'etiqueta: valor' normal.
    # La maña: agarramos TODO el bloque entre 'DATOS' y 'Señas particulares:'
    # (que es donde vive), y despues le borramos la etiqueta que quedo colgada
    # en medio. Asi recuperamos el texto completo aunque este partido.
    m_cf = re.search(
        r"DATOS\s*(.*?)\s*Señas particulares:",
        texto, re.IGNORECASE | re.DOTALL)
    if m_cf:
        cf = m_cf.group(1)
        cf = re.sub(r"Características f[ií]sicas:", " ", cf, flags=re.IGNORECASE)
        cf = re.sub(r"\s*\n\s*", " ", cf)   # unir las lineas partidas
        cf = re.sub(r"\s+", " ", cf).strip()
        d["caracteristicas_fisicas"] = cf
    else:
        d["caracteristicas_fisicas"] = ""

    # Estos dos si son bloques "normales": el texto va todo junto despues de su
    # etiqueta, asi que el helper generico los saca sin problema.
    d["senas_particulares"] = extraer_bloque_datos(
        texto, "Señas particulares:", r"Prendas de vestir:")
    d["prendas_vestir"] = extraer_bloque_datos(
        texto, "Prendas de vestir:", r"Autoridades Competentes:")

    return d


def validar_datos(datos: dict):
    """
    Revisa que el PDF tenga los datos minimos indispensables.
    Lanza ErrorPDF con un mensaje claro si falta algo critico.
    Esto evita meter registros vacios o incompletos al Excel.
    """
    faltantes = []
    if not (datos.get("nombre") or "").strip():
        faltantes.append("NOMBRE")
    if not (datos.get("fub") or "").strip():
        faltantes.append("FUB")
    if faltantes:
        raise ErrorPDF("no parece una ficha de Alerta valida; "
                       f"falta(n): {', '.join(faltantes)}")


def quitar_acentos(texto: str) -> str:
    """
    Quita acentos de las vocales (Á->A, É->E, etc.) pero CONSERVA la enie (Ñ),
    porque en nombres mexicanos la enie es una letra distinta, no un acento.
    Tambien conserva la dieresis de la u (Ü) si apareciera.
    """
    if not isinstance(texto, str):
        return texto
    # El truco de las eñes: si dejaramos que la normalizacion Unicode hiciera
    # su trabajo a secas, la Ñ tambien se descompondria y terminaria como N.
    # Y un MUÑOZ no puede volverse MUNOZ, la eñe es otra letra, no una N con
    # acento. Asi que ANTES de limpiar, escondemos las eñes (y las ü) detras de
    # unos caracteres marcadores invisibles que la normalizacion no toca...
    protegidos = (texto
                  .replace("Ñ", "\x00").replace("ñ", "\x01")
                  .replace("Ü", "\x02").replace("ü", "\x03"))
    # ...quitamos los acentos del resto con toda tranquilidad...
    sin_acentos = unicodedata.normalize("NFD", protegidos)
    sin_acentos = "".join(c for c in sin_acentos
                          if unicodedata.category(c) != "Mn")
    # ...y al final devolvemos las eñes y ü a su lugar. Listo, MUÑOZ sobrevive.
    return (sin_acentos
            .replace("\x00", "Ñ").replace("\x01", "ñ")
            .replace("\x02", "Ü").replace("\x03", "ü"))


def limpiar_suave_sql(texto: str) -> str:
    """
    Limpieza SUAVE para campos descriptivos largos (caracteristicas fisicas,
    señas particulares, prendas de vestir).

    A diferencia de limpiar_para_sql (estricta), aqui SI conservamos comas,
    dos puntos, acentos y demas puntuacion, porque en una oracion completa
    esos signos dan contexto y no se deben perder.

    Solo neutralizamos lo que puede causar problemas al insertar en SQL:
      - la comilla simple  '   (delimita cadenas en SQL)
      - el punto y coma     ;  (separa instrucciones en SQL)
    Ambos se reemplazan de forma segura. Tambien colapsa espacios.
    """
    if not isinstance(texto, str):
        return texto
    limpio = texto.replace("'", " ").replace(";", ",")
    limpio = re.sub(r"\s+", " ", limpio).strip()
    return limpio


def limpiar_para_sql(texto: str) -> str:
    """
    Deja solo caracteres compatibles con la base de datos destino:
    letras (incluida Ñ/ñ), numeros, espacio, guion '-' y diagonal '/'.
    Elimina cualquier otro simbolo (puntos, comas, parentesis, etc.).
    Tambien colapsa espacios multiples y recorta los extremos.

    Se asume que el texto ya viene SIN ACENTOS (quitar_acentos se aplica
    antes), por eso aqui solo permitimos A-Z, a-z, Ñ y ñ entre las letras.
    """
    if not isinstance(texto, str):
        return texto
    # Conserva solo lo permitido. \w incluiria '_' y acentos, por eso
    # definimos el conjunto a mano de forma explicita.
    permitido = re.sub(r"[^A-Za-z0-9Ññ /\-]", "", texto)
    # Colapsa espacios repetidos que pudieran quedar al quitar simbolos
    permitido = re.sub(r"\s+", " ", permitido).strip()
    return permitido


def a_mayusculas(datos: dict) -> dict:
    """
    Limpia los campos de texto para que sean aptos para Excel y SQL.

    Dos tratamientos distintos:
    - Campos descriptivos largos (caracteristicas fisicas, señas, prendas):
      limpieza SUAVE. Se conservan comas, acentos y puntuacion para no
      perder contexto; solo se neutraliza lo peligroso para SQL. Se
      escriben en MAYUSCULAS pero CON acentos.
    - Todos los demas campos de texto: limpieza ESTRICTA (mayusculas, sin
      acentos salvo Ñ, solo letras/numeros/espacio/-//).

    No toca las fechas (son objetos date).
    """
    # Aqui decidimos, campo por campo, que tratamiento le toca. El orden de los
    # if importa: primero descartamos lo que NO se toca, luego los casos
    # especiales, y al final el tratamiento normal (el "else").
    no_tocar = {"fecha_hecho", "fecha_percato"}          # fechas: son date, ni tocar
    descriptivos = {"caracteristicas_fisicas", "senas_particulares",
                    "prendas_vestir"}                     # oraciones: limpieza suave
    horas = {"hora_hecho", "hora_percato"}               # HH:MM: dejar los dos puntos
    salida = {}
    for k, v in datos.items():
        if not isinstance(v, str) or k in no_tocar:
            # fechas (objetos date) o cualquier cosa que no sea texto: tal cual
            salida[k] = v
        elif k in horas:
            salida[k] = v.strip()  # dejar HH:MM intacto, si le quitamos ':' se arruina
        elif k in descriptivos:
            # mayuscula + limpieza suave: conserva comas y acentos del contexto
            salida[k] = limpiar_suave_sql(v.upper())
        else:
            # el resto: mayuscula, fuera acentos (menos Ñ) y solo lo permitido
            salida[k] = limpiar_para_sql(quitar_acentos(v.upper()))
    return salida


def fila_para_nuevo_registro(ws, col_nombre):
    """
    Encuentra la fila donde escribir el nuevo registro.

    Importante: en este Excel las columnas ID y EXPEDIENTE ya vienen
    pre-llenadas hasta muchas filas adelante, aunque la fila este vacia.
    Por eso NO podemos guiarnos por el ID. El indicador de un registro
    real es que la columna NOMBRE tenga texto.

    Devuelve la primera fila cuyo NOMBRE este vacio (= primer hueco real).
    """
    fila = PRIMERA_FILA_DATOS
    while True:
        val = ws.cell(row=fila, column=col_nombre).value
        if val is None or str(val).strip() == "":
            return fila
        fila += 1


def escribir_fila(ws, datos: dict, mapa: dict, fecha_registro=None):
    """
    Escribe un registro en la primera fila sin NOMBRE de un worksheet
    YA ABIERTO, usando 'mapa' (dato -> numero de columna) que se obtuvo
    leyendo los encabezados de la fila 1. Asi el orden de las columnas
    puede cambiar sin romper nada.

    'fecha_registro' es la fecha que va en la columna FECHA DE REGISTRO.
    Si no se pasa, se usa la fecha de hoy. Sirve para cuando el correo
    llego un dia distinto al que se captura.

    No guarda el archivo (lo hace quien llama, una sola vez al final).
    Devuelve (nuevo_id, fila, valores).
    """
    col_id = buscar_columna_id(ws)
    col_nombre = mapa.get("nombre")
    if col_nombre is None:
        raise ErrorPDF("la hoja no tiene una columna 'NOMBRE'")

    fila = fila_para_nuevo_registro(ws, col_nombre)
    fila_modelo = fila - 1  # la fila de arriba, que ya tiene el formato bueno

    # --- Copiar el formato de la fila anterior a la fila nueva ---
    # openpyxl no hereda estilos en celdas vacias. Si escribimos sin mas, la
    # fila nueva sale "pelona", sin el ID en rojo, sin bordes ni la fuente
    # correcta. La solucion: clonar celda por celda el estilo de la fila de
    # arriba (que ya esta bien formateada). copy() hace copia independiente,
    # porque si asignaramos el estilo directo quedarian ligados y cambiar uno
    # cambiaria el otro.
    if fila_modelo >= PRIMERA_FILA_DATOS:
        for col in range(1, ws.max_column + 1):
            origen = ws.cell(row=fila_modelo, column=col)
            destino = ws.cell(row=fila, column=col)
            if origen.has_style:
                destino.font = copy(origen.font)
                destino.fill = copy(origen.fill)
                destino.border = copy(origen.border)
                destino.alignment = copy(origen.alignment)
                destino.number_format = origen.number_format
                destino.protection = copy(origen.protection)
        if fila_modelo in ws.row_dimensions and ws.row_dimensions[fila_modelo].height:
            ws.row_dimensions[fila].height = ws.row_dimensions[fila_modelo].height

    # El ID en este Excel ya viene pre-llenado hasta muy abajo. Asi que lo
    # normal es RESPETAR el que ya esta. Solo si por alguna razon estuviera
    # vacio, lo calculamos sumandole 1 al de la fila de arriba.
    id_actual = ws.cell(row=fila, column=col_id).value
    if id_actual is None or str(id_actual).strip() == "":
        id_arriba = ws.cell(row=fila_modelo, column=col_id).value
        nuevo_id = int(id_arriba) + 1 if id_arriba not in (None, "") else 1
        ws.cell(row=fila, column=col_id, value=nuevo_id)
    else:
        nuevo_id = int(id_actual)

    # Mismo criterio con el EXPEDIENTE: si ya esta puesto, no lo pisamos. Si
    # esta vacio, lo generamos con el formato AN/{id}/{año}.
    col_exp = mapa.get("expediente")
    expediente = ""
    if col_exp:
        exp_actual = ws.cell(row=fila, column=col_exp).value
        if exp_actual is None or str(exp_actual).strip() == "":
            ws.cell(row=fila, column=col_exp,
                    value=f"AN/{nuevo_id}/{ANIO_EXPEDIENTE}")
        expediente = ws.cell(row=fila, column=col_exp).value

    # Juntamos todo lo que vamos a escribir: lo que sacamos del PDF, mas la
    # fecha de registro (la que eligio el usuario, o hoy si no eligio) y el
    # expediente que acabamos de resolver.
    valores = dict(datos)  # copia de lo extraido del PDF
    valores["fecha_registro"] = fecha_registro or dt.date.today()
    valores["expediente"] = expediente

    numericos = {"edad_desaparecer", "edad_actual"}

    # Escribir cada dato en SU columna segun el mapa (por nombre).
    # Si la hoja no tiene ese encabezado, simplemente se omite ese dato.
    for dato, col in mapa.items():
        if dato in ("nombre",):
            valor = valores.get("nombre")
        else:
            valor = valores.get(dato)
        if valor is None:
            continue
        if dato in numericos:
            try:
                valor = int(valor)
            except (ValueError, TypeError):
                pass
        ws.cell(row=fila, column=col, value=valor)

    return nuevo_id, fila, valores


def leer_fubs_existentes(ws, col_fub, col_nombre) -> set:
    """
    Lee todos los FUB que ya estan registrados y los devuelve en un
    conjunto normalizado (mayusculas, sin espacios) para comparar.
    Sirve para detectar duplicados.
    """
    fubs = set()
    if col_fub is None or col_nombre is None:
        return fubs
    fila = PRIMERA_FILA_DATOS
    while True:
        nombre = ws.cell(row=fila, column=col_nombre).value
        if nombre is None or str(nombre).strip() == "":
            break  # se acabaron los registros reales
        fub = ws.cell(row=fila, column=col_fub).value
        if fub is not None and str(fub).strip():
            fubs.add(str(fub).strip().upper())
        fila += 1
    return fubs


def listar_hojas(ruta_excel) -> list:
    """
    Devuelve la lista de nombres de hojas de un archivo Excel.
    La interfaz la usa para llenar el menu desplegable de seleccion de hoja.
    Si el archivo no se puede abrir, devuelve lista vacia.
    """
    try:
        wb = openpyxl.load_workbook(ruta_excel, read_only=True)
        nombres = list(wb.sheetnames)
        wb.close()
        return nombres
    except Exception:
        return []


def procesar_carpeta(carpeta_pdfs, ruta_excel, carpeta_procesados,
                     hoja=None, hojas=None, config=None, fecha_registro=None,
                     carpeta_fotos=None, log=print):
    """
    Procesa todos los PDF de 'carpeta_pdfs' y los registra en UNA o VARIAS
    hojas del Excel en una sola pasada.

    Parametros de hoja (usar uno):
      - hojas: lista de nombres de hoja donde registrar (ej. ['Hoja1','Hoja2'])
      - hoja:  un solo nombre (compatibilidad con la version anterior)

    Comportamiento:
      - Cada PDF se lee UNA sola vez.
      - Se escribe en cada hoja activa donde su FUB no exista todavia
        (duplicados independientes POR HOJA).
      - El PDF se mueve UNA vez segun el resultado global:
          * a 'procesados'  si se agrego al menos en una hoja
          * a 'duplicados'  si ya existia en todas las hojas activas
          * a 'errores'     si fallo la lectura/validacion
      - Las subcarpetas 'duplicados'/'errores' se crean solo si hacen falta.

    Devuelve un resumen: dict con 'agregados', 'duplicados' y 'errores'
    (esta ultima con tuplas (nombre_pdf, motivo)).
    """
    carpeta_pdfs = Path(carpeta_pdfs)
    ruta_excel = Path(ruta_excel)
    carpeta_procesados = Path(carpeta_procesados)

    resumen = {"agregados": [], "duplicados": [], "errores": []}

    # Normalizar la lista de hojas a usar
    if hojas is None:
        hojas = [hoja if hoja else HOJA_DEFAULT]
    hojas = [h for h in hojas if h]  # quitar vacios
    if not hojas:
        log("ERROR: no se indico ninguna hoja donde registrar.")
        return resumen

    # --- Validaciones basicas de entrada ---
    if not carpeta_pdfs.is_dir():
        log(f"ERROR: la carpeta de PDFs no existe: {carpeta_pdfs}")
        return resumen
    if not ruta_excel.is_file():
        log(f"ERROR: no encuentro el Excel: {ruta_excel}")
        return resumen

    pdfs = sorted(carpeta_pdfs.glob("*.pdf"))
    if not pdfs:
        log("No hay archivos PDF en la carpeta.")
        return resumen

    # Si el usuario pidio guardar fotos pero PyMuPDF no esta disponible, lo
    # avisamos claro (antes se quedaba callado y parecia que no hacia nada).
    if carpeta_fotos and not _HAY_FITZ:
        log("AVISO: se pidio guardar fotos, pero la libreria para leer "
            "imagenes de PDF (PyMuPDF) no esta instalada. Los registros se "
            "haran normal, pero NO se guardaran fotos. Instala con: "
            "pip install pymupdf")

    # --- Mapa de encabezados: desde la config del usuario si se paso ---
    mapa_encabezados = None
    if config is not None:
        try:
            import config_campos
            mapa_encabezados = config_campos.mapa_encabezados_desde_config(config)
        except Exception:
            mapa_encabezados = None  # ante cualquier fallo, usa el fijo

    # --- Respaldo de seguridad del Excel ---
    # Antes de tocar el Excel, hacemos una copia .bak. Si algo sale mal a mitad
    # del proceso, el usuario tiene de donde recuperar su base. Si ni el
    # respaldo se puede crear, mejor ni empezamos: cancelamos por seguridad.
    try:
        respaldo = ruta_excel.with_suffix(".bak.xlsx")
        shutil.copy(ruta_excel, respaldo)
        log(f"Respaldo creado: {respaldo.name}")
    except Exception as e:
        log(f"ERROR: no se pudo crear el respaldo del Excel ({e}). "
            "Se cancela el proceso por seguridad.")
        return resumen

    # --- Abrir el Excel una sola vez ---
    # Lo abrimos UNA vez para toda la tanda (no por cada PDF), es mas rapido.
    # El PermissionError casi siempre significa que el usuario dejo el Excel
    # abierto; se lo decimos claro en vez de soltarle un error tecnico.
    try:
        wb = openpyxl.load_workbook(ruta_excel)
    except PermissionError:
        log("ERROR: el Excel esta abierto. Cierralo y vuelve a intentar.")
        return resumen
    except Exception as e:
        log(f"ERROR: no se pudo abrir el Excel ({type(e).__name__}: {e}).")
        return resumen

    # --- Preparar cada hoja destino ---
    # Por cada hoja donde vamos a escribir, preparamos de una vez todo lo que
    # necesitaremos en el bucle: su worksheet, el mapa de columnas (por nombre
    # de encabezado) y el set de FUBs que ya tiene (para detectar duplicados).
    # Lo guardamos en 'destinos' para no recalcularlo con cada PDF.
    destinos = []
    for nombre_hoja in hojas:
        if nombre_hoja not in wb.sheetnames:
            log(f"ERROR: el Excel no tiene la hoja '{nombre_hoja}'. "
                f"Hojas disponibles: {', '.join(wb.sheetnames)}.")
            return resumen
        ws = wb[nombre_hoja]
        mapa = mapear_columnas(ws, mapa_encabezados)
        # Sin columna NOMBRE no sabriamos donde empieza el hueco libre ni donde
        # escribir. Es requisito minimo de cualquier hoja destino.
        if "nombre" not in mapa:
            log(f"ERROR: la hoja '{nombre_hoja}' no tiene columna 'NOMBRE'. "
                "No se puede saber donde escribir.")
            return resumen
        fubs = leer_fubs_existentes(ws, mapa.get("fub"), mapa.get("nombre"))
        destinos.append({
            "nombre": nombre_hoja, "ws": ws, "mapa": mapa, "fubs": fubs,
        })
        log(f"Hoja '{nombre_hoja}': {len(fubs)} FUB unicos ya registrados.")

    # --- Asegurar la carpeta de procesados ---
    try:
        carpeta_procesados.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        log(f"ERROR: no se pudo crear la carpeta de procesados ({e}).")
        return resumen

    # Las subcarpetas de duplicados y errores NO se crean aqui a proposito.
    # Solo se crean si de verdad aparece un duplicado o un error (ver el
    # _asegurar_carpeta dentro del bucle). Asi no ensuciamos con carpetas vacias.
    carpeta_duplicados = carpeta_procesados / "duplicados"
    carpeta_errores = carpeta_procesados / "errores"

    log("-" * 50)
    hubo_cambios = False

    for pdf in pdfs:
        # Cada PDF se lee UNA sola vez, aunque vaya a varias hojas. Todo el
        # trabajo de un PDF va envuelto en try/except: si uno falla, lo
        # anotamos y seguimos con los demas, no se cae toda la tanda.
        try:
            texto = leer_texto_pdf(str(pdf))
            datos = a_mayusculas(extraer_datos(texto, config))
            validar_datos(datos)
            fub = (datos.get("fub") or "").strip().upper()

            # Llevamos la cuenta de en que hojas se escribio y en cuales ya
            # existia. Con esto decidimos despues a donde mover el PDF.
            agregado_en = []   # hojas donde se escribio
            duplicado_en = []  # hojas donde ya existia

            # Recorremos cada hoja destino por separado. El duplicado es POR
            # HOJA: puede que ya este en la completa pero no en la
            # personalizada, y entonces se agrega solo donde falta.
            for d in destinos:
                if fub in d["fubs"]:
                    duplicado_en.append(d["nombre"])
                else:
                    nuevo_id, fila, _ = escribir_fila(d["ws"], datos, d["mapa"],
                                                      fecha_registro)
                    # Agregamos el fub al set en memoria para que, si el mismo
                    # PDF viene dos veces en esta tanda, el segundo ya lo cache.
                    d["fubs"].add(fub)
                    agregado_en.append((d["nombre"], nuevo_id, fila))
                    hubo_cambios = True

            # --- Guardar la foto de la persona (si se pidio) ---
            # Se guarda siempre que el PDF sea valido y tenga FUB, sin importar
            # si termina como nuevo o duplicado: la foto es util igual.
            if carpeta_fotos and fub and _HAY_FITZ:
                if guardar_foto(str(pdf), fub, carpeta_fotos, log):
                    log(f"      foto guardada: {fub}.jpg")
                else:
                    log(f"      (no se encontro foto en {pdf.name})")

            # Segun el resultado, el PDF va a una carpeta u otra:
            if agregado_en:
                # Se escribio al menos en una hoja -> cuenta como agregado.
                detalle = ", ".join(f"{h} (ID {i}, fila {f})"
                                    for h, i, f in agregado_en)
                extra = ""
                if duplicado_en:
                    extra = f"  [ya existia en: {', '.join(duplicado_en)}]"
                log(f"  [+] {pdf.name}: agregado en {detalle}.{extra}")
                resumen["agregados"].append(pdf.name)
                _mover(pdf, carpeta_procesados, log)
            else:
                # No entro en ninguna: ya estaba en todas. Va a 'duplicados'.
                log(f"  [=] {pdf.name}: duplicado en todas las hojas "
                    f"({', '.join(duplicado_en)}). Se omite.")
                resumen["duplicados"].append(pdf.name)
                _asegurar_carpeta(carpeta_duplicados)
                _mover(pdf, carpeta_duplicados, log)

        except ErrorPDF as e:
            # Error "esperado" y con mensaje claro (PDF dañado, sin datos...).
            log(f"  [x] {pdf.name}: ERROR -> {e}")
            resumen["errores"].append((pdf.name, str(e)))
            _asegurar_carpeta(carpeta_errores)
            _mover(pdf, carpeta_errores, log)

        except Exception as e:
            # Red de seguridad: cualquier otro error raro que no previmos.
            # No se pierde, se anota con su tipo y el PDF va a 'errores'.
            motivo = f"error inesperado ({type(e).__name__}: {e})"
            log(f"  [x] {pdf.name}: {motivo}")
            resumen["errores"].append((pdf.name, motivo))
            _asegurar_carpeta(carpeta_errores)
            _mover(pdf, carpeta_errores, log)

    # --- Guardar el Excel una sola vez al final ---
    if hubo_cambios:
        try:
            wb.save(ruta_excel)
        except PermissionError:
            log("ERROR al guardar: el Excel esta abierto. "
                "Cierralo; los cambios de esta tanda NO se guardaron.")
            return resumen
        except Exception as e:
            log(f"ERROR al guardar el Excel ({type(e).__name__}: {e}). "
                "Los cambios podrian no haberse guardado.")
            return resumen

    # --- Resumen final ---
    log("-" * 50)
    log(f"Listo. Agregados con exito: {len(resumen['agregados'])} | "
        f"Duplicados: {len(resumen['duplicados'])} | "
        f"Con error: {len(resumen['errores'])}.")

    if resumen["errores"]:
        log("")
        log("Archivos que NO se pudieron procesar:")
        for nombre, motivo in resumen["errores"]:
            log(f"   - {nombre}: {motivo}")

    return resumen


def _asegurar_carpeta(carpeta: Path):
    """Crea la carpeta solo si no existe. Se llama justo antes de necesitarla,
    para que las subcarpetas de duplicados/errores aparezcan unicamente
    cuando de verdad hay algo que mover ahi."""
    if not carpeta.exists():
        carpeta.mkdir(parents=True, exist_ok=True)


def _mover(pdf: Path, destino: Path, log):
    """Mueve un PDF a la carpeta destino. Si ya existe alli, le agrega
    un sufijo para no sobrescribir."""
    # Si en la carpeta destino ya hay un archivo con el mismo nombre (ej.
    # reprocesaste el mismo PDF otro dia), NO lo pisamos. Le buscamos un nombre
    # libre agregando _1, _2, etc. Asi nunca se pierde un archivo por choque
    # de nombres.
    objetivo = destino / pdf.name
    if objetivo.exists():
        i = 1
        while True:
            cand = destino / f"{pdf.stem}_{i}{pdf.suffix}"
            if not cand.exists():
                objetivo = cand
                break
            i += 1
    try:
        shutil.move(str(pdf), str(objetivo))
    except Exception as e:
        log(f"      (no se pudo mover {pdf.name}: {e})")


# ---------------------------------------------------------------------------
# CESE DE DIFUSION
# Marca personas como LOCALIZADAS: pone STATUS=LOCALIZADO, la fecha de
# localizado, y pinta de verde las celdas con texto de esa fila.
# ---------------------------------------------------------------------------

# Verde claro tipo Excel (el que se ve en la paleta estandar de Office).
VERDE_LOCALIZADO = "A9D08E"


def _construir_indice_fub(ws, col_fub, col_nombre) -> dict:
    """
    Recorre la hoja y arma un indice {fub_normalizado: fila} para poder
    encontrar rapido en que fila esta cada persona por su FUB.
    Recorremos mientras haya NOMBRE (ese es el indicador de registro real).
    """
    indice = {}
    if col_fub is None or col_nombre is None:
        return indice
    fila = PRIMERA_FILA_DATOS
    while True:
        nombre = ws.cell(row=fila, column=col_nombre).value
        if nombre is None or str(nombre).strip() == "":
            break
        fub = ws.cell(row=fila, column=col_fub).value
        if fub is not None and str(fub).strip():
            indice[str(fub).strip().upper()] = fila
        fila += 1
    return indice


def marcar_localizados(ruta_excel, fubs, hoja=None, fecha_localizado=None,
                       log=print):
    """
    Marca como LOCALIZADAS a las personas cuyos FUB se pasan en 'fubs'.

    Para cada FUB encontrado en la hoja:
      - escribe 'LOCALIZADO' en la columna STATUS
      - escribe la fecha en la columna FECHA DE LOCALIZADO
      - pinta de verde SOLO las celdas de esa fila que tienen texto

    Devuelve un resumen: {'localizados': [...], 'no_encontrados': [...]}.
    'fubs' es una lista de folios (se normalizan a mayusculas sin espacios).
    """
    resumen = {"localizados": [], "no_encontrados": []}
    ruta_excel = Path(ruta_excel)

    if not ruta_excel.is_file():
        log(f"ERROR: no encuentro el Excel: {ruta_excel}")
        return resumen

    # Limpiar y quitar repetidos de la lista de FUBs recibida
    fubs_limpios = []
    vistos = set()
    for f in fubs:
        f = (f or "").strip().upper()
        if f and f not in vistos:
            vistos.add(f)
            fubs_limpios.append(f)
    if not fubs_limpios:
        log("No se indico ningun FUB para localizar.")
        return resumen

    # Respaldo antes de tocar nada
    try:
        respaldo = ruta_excel.with_suffix(".bak.xlsx")
        shutil.copy(ruta_excel, respaldo)
        log(f"Respaldo creado: {respaldo.name}")
    except Exception as e:
        log(f"ERROR: no se pudo crear el respaldo ({e}). Se cancela.")
        return resumen

    # Abrir Excel
    try:
        wb = openpyxl.load_workbook(ruta_excel)
    except PermissionError:
        log("ERROR: el Excel esta abierto. Cierralo y vuelve a intentar.")
        return resumen
    except Exception as e:
        log(f"ERROR: no se pudo abrir el Excel ({type(e).__name__}: {e}).")
        return resumen

    hoja_usar = hoja if hoja else HOJA_DEFAULT
    if hoja_usar not in wb.sheetnames:
        log(f"ERROR: el Excel no tiene la hoja '{hoja_usar}'. "
            f"Hojas disponibles: {', '.join(wb.sheetnames)}.")
        return resumen
    ws = wb[hoja_usar]

    # Ubicar columnas por su encabezado (igual que en el resto del programa)
    mapa = mapear_columnas(ws)
    col_fub = mapa.get("fub")
    col_nombre = mapa.get("nombre")
    col_status = _buscar_columna_por_nombre(ws, ["status", "estatus", "estado"])
    col_fecha_loc = _buscar_columna_por_nombre(
        ws, ["fecha de localizado", "fecha localizado"])

    # Sin estas columnas no podemos trabajar; avisamos claro cual falta.
    faltan = []
    if col_fub is None:
        faltan.append("FUB")
    if col_status is None:
        faltan.append("STATUS")
    if col_fecha_loc is None:
        faltan.append("FECHA DE LOCALIZADO")
    if faltan:
        log(f"ERROR: la hoja '{hoja_usar}' no tiene la(s) columna(s): "
            f"{', '.join(faltan)}. No se puede continuar.")
        return resumen

    fecha_localizado = fecha_localizado or dt.date.today()
    relleno = PatternFill(start_color=VERDE_LOCALIZADO,
                          end_color=VERDE_LOCALIZADO, fill_type="solid")

    # Armar el indice FUB -> fila una sola vez (rapido aunque haya miles)
    indice = _construir_indice_fub(ws, col_fub, col_nombre)

    hubo_cambios = False
    for fub in fubs_limpios:
        fila = indice.get(fub)
        if fila is None:
            log(f"  [?] FUB {fub}: no se encontro en la hoja.")
            resumen["no_encontrados"].append(fub)
            continue

        # Escribir STATUS y FECHA DE LOCALIZADO
        ws.cell(row=fila, column=col_status, value="LOCALIZADO")
        ws.cell(row=fila, column=col_fecha_loc, value=fecha_localizado)

        # Pintar de verde SOLO las celdas de la fila que tienen texto/valor
        for col in range(1, ws.max_column + 1):
            celda = ws.cell(row=fila, column=col)
            if celda.value is not None and str(celda.value).strip() != "":
                celda.fill = relleno

        nombre = ws.cell(row=fila, column=col_nombre).value if col_nombre else ""
        log(f"  [+] FUB {fub}: LOCALIZADO en fila {fila} ({nombre}).")
        resumen["localizados"].append(fub)
        hubo_cambios = True

    # Guardar una sola vez
    if hubo_cambios:
        try:
            wb.save(ruta_excel)
        except PermissionError:
            log("ERROR al guardar: el Excel esta abierto. Los cambios NO se "
                "guardaron.")
            return resumen
        except Exception as e:
            log(f"ERROR al guardar ({type(e).__name__}: {e}).")
            return resumen

    log("-" * 50)
    log(f"Listo. Localizados: {len(resumen['localizados'])} | "
        f"No encontrados: {len(resumen['no_encontrados'])}.")
    if resumen["no_encontrados"]:
        log("")
        log("FUB que no se encontraron en la hoja:")
        for f in resumen["no_encontrados"]:
            log(f"   - {f}")
    return resumen


def _buscar_columna_por_nombre(ws, nombres_posibles):
    """
    Busca en la fila 1 una columna cuyo encabezado (normalizado) coincida con
    alguno de los nombres posibles. Devuelve el numero de columna o None.
    """
    encabezados = {}
    for col in range(1, ws.max_column + 1):
        norm = _normalizar_encabezado(ws.cell(row=1, column=col).value)
        if norm:
            encabezados[norm] = col
    for nombre in nombres_posibles:
        clave = _normalizar_encabezado(nombre)
        if clave in encabezados:
            return encabezados[clave]
    return None


def main():
    args = sys.argv[1:]

    # Modo carpeta:
    #   python capturar_alerta.py --carpeta "ruta/pdfs" --excel "base.xlsx"
    #                             --procesados "ruta/procesados"
    if "--carpeta" in args:
        carpeta = args[args.index("--carpeta") + 1]
        ruta_excel = EXCEL_DEFAULT
        if "--excel" in args:
            ruta_excel = args[args.index("--excel") + 1]
        if "--procesados" in args:
            procesados = args[args.index("--procesados") + 1]
        else:
            procesados = str(Path(carpeta) / "procesados")
        hoja = None
        if "--hoja" in args:
            hoja = args[args.index("--hoja") + 1]
        procesar_carpeta(carpeta, ruta_excel, procesados, hoja=hoja)
        return

    # Modo un solo PDF (compatibilidad con lo anterior):
    #   python capturar_alerta.py "flyer.pdf" [--excel "base.xlsx"]
    if not args:
        print('Uso:')
        print('  Un PDF:   python capturar_alerta.py "flyer.pdf" [--excel "base.xlsx"]')
        print('  Carpeta:  python capturar_alerta.py --carpeta "pdfs" '
              '--excel "base.xlsx" --procesados "procesados" [--hoja "Hoja1"]')
        sys.exit(1)

    ruta_pdf = args[0]
    ruta_excel = EXCEL_DEFAULT
    if "--excel" in args:
        ruta_excel = args[args.index("--excel") + 1]
    hoja_cli = args[args.index("--hoja") + 1] if "--hoja" in args else HOJA_DEFAULT

    if not Path(ruta_pdf).exists():
        print(f"No encuentro el PDF: {ruta_pdf}")
        sys.exit(1)
    if not Path(ruta_excel).exists():
        print(f"No encuentro el Excel: {ruta_excel}")
        sys.exit(1)

    respaldo = Path(ruta_excel).with_suffix(".bak.xlsx")
    shutil.copy(ruta_excel, respaldo)

    try:
        wb = openpyxl.load_workbook(ruta_excel)
    except PermissionError:
        print("ERROR: el Excel esta abierto. Cierralo y vuelve a intentar.")
        sys.exit(1)
    if hoja_cli not in wb.sheetnames:
        print(f"ERROR: el Excel no tiene la hoja '{hoja_cli}'. "
              f"Hojas disponibles: {', '.join(wb.sheetnames)}.")
        sys.exit(1)
    ws = wb[hoja_cli]
    mapa = mapear_columnas(ws)

    try:
        texto = leer_texto_pdf(ruta_pdf)
        datos = a_mayusculas(extraer_datos(texto))
        validar_datos(datos)
    except ErrorPDF as e:
        print(f"\nERROR: {e}")
        return

    # Aviso de duplicado tambien en modo un PDF
    fub = (datos.get("fub") or "").strip().upper()
    if fub and fub in leer_fubs_existentes(ws, mapa.get("fub"), mapa.get("nombre")):
        print(f"\n[=] Duplicado: el FUB {fub} ya esta registrado. No se agrega.")
        return

    nuevo_id, fila, valores = escribir_fila(ws, datos, mapa)
    wb.save(ruta_excel)

    print(f"\nRegistro agregado. ID {nuevo_id} en la fila {fila}.")
    print(f"Respaldo previo guardado en: {respaldo.name}\n")
    print("Datos capturados:")
    for k, v in valores.items():
        print(f"  {k:18s}: {v}")
    print("\nRevisa siempre el registro antes de darlo por bueno.")


if __name__ == "__main__":
    main()